from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from config import config_manager  # config_manager에서 경로 설정 로드
from analysis.plotter import Plotter

class ExcelFormatter:
    def __init__(self, title="Financial Data", zoom_scale=60, output_file=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = title
        self.zoom_scale = zoom_scale

        # 설정된 경로에서 파일명을 동적으로 설정
        config = config_manager.load_config()
        if output_file is None:
            self.output_file = f"{config['csv_save_path']}/financial_analysis_results.xlsx"
        else:
            self.output_file = output_file

        # 스타일 초기화
        self.medium_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.default_font = Font(name="맑은 고딕", bold=False)
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # 초기 기본값 설정
        self._initialize_defaults()

    def _initialize_defaults(self):
        """전체 시트 초기화"""
        self.ws.sheet_view.zoomScale = self.zoom_scale
        for col_idx in range(1, 45):
            self.ws.column_dimensions[get_column_letter(col_idx)].width = 8.1
        for row_idx in range(1, 1000):
            self.ws.row_dimensions[row_idx].height = 22
        for row in self.ws.iter_rows(min_row=1, max_row=1000, min_col=1, max_col=182):
            for cell in row:
                cell.fill = self.white_fill

    def apply_color_fill(self, color_fill_settings):
        """
        지정된 영역에 색상을 채우는 메서드.
        :param color_fill_settings: 색상을 적용할 영역과 색상 코드 리스트
        """
        for cell_range, color in color_fill_settings:
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for row in self.ws[cell_range]:
                for cell in row:
                    cell.fill = fill

    def create_section(self, start_row, company_name, income_data, balance_data):
        """
        특정 시작 행에서 데이터를 삽입하며 섹션을 생성합니다.
        :param start_row: 데이터 시작 행
        :param company_name: 회사명
        :param income_data: 손익계산서 데이터 (DataFrame)
        :param balance_data: 재무상태표 데이터 (DataFrame)
        """
        # JSON에서 항목 코드 매핑 로드
        item_codes = config_manager.load_item_codes()
        INCOME_STATEMENT_ITEM_CODES = item_codes["INCOME_STATEMENT_ITEM_CODES"]
        BALANCE_SHEET_ITEM_CODES = item_codes["BALANCE_SHEET_ITEM_CODES"]

        # 삽입할 항목
        target_items = {
            "ifrs-full_ProfitLoss": "당기순이익",
            "ifrs-full_CostOfSales": "매출원가",
            "dart_OperatingIncomeLoss": "영업이익",
            "ifrs-full_Liabilities": "부채비율",
            "ifrs-full_CurrentAssets": "유동비율",
        }
        # 테두리 설정
        specified_ranges = [
            f"B{start_row + 3}:I{start_row + 9}", f"K{start_row + 1}:W{start_row + 12}",
            f"Y{start_row + 1}:AR{start_row + 12}",
            f"Q{start_row + 3}:Q{start_row + 12}", f"R{start_row + 4}:W{start_row + 12}",
            f"Y{start_row + 4}:AD{start_row + 12}", f"AE{start_row + 3}:AE{start_row + 12}",
            f"AF{start_row + 4}:AK{start_row + 12}", f"AL{start_row + 3}:AL{start_row + 12}",
            f"AM{start_row + 4}:AR{start_row + 12}"
        ]
        self.apply_border(specified_ranges)

        # 셀 병합 및 텍스트 설정
        merge_and_set = [
            (f"B{start_row + 3}:C{start_row + 4}", company_name), (f"D{start_row + 3}:D{start_row + 4}", "2019년"),
            (f"E{start_row + 3}:E{start_row + 4}", "2020년"), (f"F{start_row + 3}:F{start_row + 4}", "2021년"),
            (f"G{start_row + 3}:G{start_row + 4}", "2022년"), (f"H{start_row + 3}:H{start_row + 4}", "2023년"),
            (f"I{start_row + 3}:I{start_row + 4}", "2024(예측)"),
            (f"B{start_row + 5}:C{start_row + 5}", "당기순이익"), (f"B{start_row + 6}:C{start_row + 6}", "매출원가"),
            (f"B{start_row + 7}:C{start_row + 7}", "영업이익"), (f"B{start_row + 8}:C{start_row + 8}", "부채비율"),
            (f"B{start_row + 9}:C{start_row + 9}", "유동비율"),
            (f"K{start_row + 1}:W{start_row + 2}", "재무상태표"),
            (f"K{start_row + 3}:P{start_row + 3}", "부채비율"), (f"R{start_row + 3}:W{start_row + 3}", "유동비율"),
            (f"Y{start_row + 1}:AR{start_row + 2}", "당기순이익"), (f"Y{start_row + 3}:AD{start_row + 3}", "당기순이익"),
            (f"AF{start_row + 3}:AK{start_row + 3}", "당기순이익"), (f"AM{start_row + 3}:AR{start_row + 3}", "영업이익")
        ]
        self.merge_cells_and_set_text(merge_and_set)

        # 바깥쪽 테두리 설정
        outer_border_ranges = [
            f"K{start_row + 3}:P{start_row + 12}", f"Q{start_row + 2}:Q{start_row + 12}",
            f"R{start_row + 3}:W{start_row + 12}", f"Y{start_row + 3}:AD{start_row + 12}",
            f"AE{start_row + 2}:AE{start_row + 12}", f"AF{start_row + 3}:AK{start_row + 12}",
            f"AL{start_row + 2}:AL{start_row + 12}", f"AM{start_row + 3}:AR{start_row + 12}"
        ]
        for cell_range in outer_border_ranges:
            self.set_outer_border(cell_range)
        # 색상 채우기 설정
        color_fill_settings = [
            (f"B{start_row + 3}:C{start_row + 4}", "FFFFD966"), (f"D{start_row + 3}:I{start_row + 4}", "FFFFE699"),
            (f"K{start_row + 3}:P{start_row + 3}", "FFC9C9C9"), (f"R{start_row + 3}:W{start_row + 3}", "FFC9C9C9"),
            (f"Y{start_row + 3}:AD{start_row + 3}", "FFC9C9C9"), (f"AF{start_row + 3}:AK{start_row + 3}", "FFC9C9C9"),
            (f"AM{start_row + 3}:AR{start_row + 3}", "FFC9C9C9"),
            (f"B{start_row + 6}:I{start_row + 6}", "FFD9D9D9"), (f"B{start_row + 8}:I{start_row + 8}", "FFD9D9D9"),
            (f"K{start_row + 1}:W{start_row + 2}", "FFFFE699"), (f"Y{start_row + 1}:AR{start_row + 2}", "FFFFE699")
        ]
        self.apply_color_fill(color_fill_settings)

        # 테이블 항목 텍스트 설정
        table_items = [
            (f"B{start_row + 5}", "당기순이익"),
            (f"B{start_row + 6}", "매출원가"),
            (f"B{start_row + 7}", "영업이익"),
            (f"B{start_row + 8}", "부채비율"),
            (f"B{start_row + 9}", "유동비율"),
        ]
        for cell, label in table_items:
            self.ws[cell].value = label
            self.ws[cell].alignment = self.center_alignment

        # 표의 행 번호와 항목 매핑
        row_mapping = {
            "당기순이익": start_row + 5,
            "매출원가": start_row + 6,
            "영업이익": start_row + 7,
            "부채비율": start_row + 8,
            "유동비율": start_row + 9,
        }

        # 손익계산서 데이터 삽입
        for idx, year in enumerate(range(2019, 2024), start=4):  # D열부터 I열까지
            income_row = income_data[income_data["연도"] == year]
            if not income_row.empty:
                for code, description in INCOME_STATEMENT_ITEM_CODES.items():
                    if code in target_items.keys():
                        filtered_row = income_row[income_row["항목코드"] == code]
                        if not filtered_row.empty:  # 데이터가 있는 경우에만 값을 가져옴
                            value = filtered_row["당기"].values[0]
                            target_row = row_mapping[target_items[code]]
                            self.ws.cell(row=target_row, column=idx).value = value
            # 2024년 예측 데이터 삽입
            if year == 2023:  # 마지막 연도까지 데이터를 기준으로 예측
                for code, description in INCOME_STATEMENT_ITEM_CODES.items():
                    if code in target_items.keys():
                        # 2024년 예측 값 계산
                        values = income_data[income_data["항목코드"] == code]["당기"].values
                        years = income_data[income_data["항목코드"] == code]["연도"].values
                        if len(values) > 1:
                            predicted_value = Plotter.predict_with_linear_regression(years, values, 2024)
                            target_row = row_mapping[target_items[code]]
                            self.ws.cell(row=target_row, column=idx + 1).value = predicted_value


        # 재무상태표 데이터 삽입
        for idx, year in enumerate(range(2019, 2024), start=4):  # D열부터 I열까지
            balance_row = balance_data[balance_data["연도"] == year]
            if not balance_row.empty:
                for code, description in BALANCE_SHEET_ITEM_CODES.items():
                    if code in target_items.keys():
                        filtered_row = balance_row[balance_row["항목코드"] == code]
                        if not filtered_row.empty:  # 데이터가 있는 경우에만 값을 가져옴
                            value = filtered_row["당기"].values[0]
                            target_row = row_mapping[target_items[code]]
                            self.ws.cell(row=target_row, column=idx).value = value
            # 2024년 예측 데이터 삽입
            if year == 2023:  # 마지막 연도까지 데이터를 기준으로 예측
                for code, description in BALANCE_SHEET_ITEM_CODES.items():
                    if code in target_items.keys():
                        # 2024년 예측 값 계산
                        values = balance_data[balance_data["항목코드"] == code]["당기"].values
                        years = balance_data[balance_data["항목코드"] == code]["연도"].values
                        if len(values) > 1:
                            predicted_value = Plotter.predict_with_linear_regression(years, values, 2024)
                            target_row = row_mapping[target_items[code]]
                            self.ws.cell(row=target_row, column=idx + 1).value = predicted_value




    def get_target_row_by_description(self, description, start_row):
        """설명을 기준으로 타겟 행 번호 반환"""
        description_mapping = {
            "당기순이익": start_row + 5,
            "매출원가": start_row + 6,
            "영업이익": start_row + 7,
            "부채총계": start_row + 8,
            "유동자산": start_row + 9
        }
        return description_mapping.get(description, start_row + 10)

    def merge_cells_and_set_text(self, merge_and_set):
        """셀 병합 및 텍스트 설정"""
        for cell_range, value in merge_and_set:
            self.ws.merge_cells(cell_range)
            start_cell = self.ws[cell_range.split(":")[0]]
            start_cell.value = value
            start_cell.alignment = self.center_alignment
            start_cell.font = self.default_font

    def apply_border(self, specified_ranges):
        """지정된 영역에 테두리를 설정"""
        for cell_range in specified_ranges:
            for row in self.ws[cell_range]:
                for cell in row:
                    cell.border = self.medium_border

    def set_outer_border(self, cell_range):
        """지정된 셀 범위에 바깥쪽 테두리를 설정"""
        rows = list(self.ws[cell_range])
        for row_idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                cell.border = Border(
                    left=self.medium_border.left if col_idx == 0 else None,
                    right=self.medium_border.right if col_idx == len(row) - 1 else None,
                    top=self.medium_border.top if row_idx == 0 else None,
                    bottom=self.medium_border.bottom if row_idx == len(rows) - 1 else None
                )



    def save_file(self):
        """파일 저장"""
        self.wb.save(self.output_file)
        print(f"[INFO] Excel file saved as {self.output_file}")
