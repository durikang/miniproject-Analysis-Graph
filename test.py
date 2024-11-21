import pandas as pd
from openpyxl import Workbook
from excel_formatter import ExcelFormatter  # 위에서 작성한 ExcelFormatter 클래스
from data.data_manager import prepare_data_for_analysis
from analysis.plotter import Plotter

# 테스트용 데이터 생성
def generate_test_data():
    """테스트용 데이터를 생성하여 CSV 파일로 저장"""
    test_data = {
        "회사명": ["A사", "A사", "B사", "B사"],
        "항목코드": ["ifrs-full_Revenue", "ifrs-full_Profit", "ifrs-full_Revenue", "ifrs-full_Profit"],
        "연도": [2019, 2020, 2019, 2020],
        "당기": [1000000, 150000, 2000000, 300000],
        "데이터종류": ["손익계산서", "손익계산서", "손익계산서", "손익계산서"]
    }
    df = pd.DataFrame(test_data)
    df.to_csv("test_income_data.csv", index=False, encoding="utf-8-sig")
    print("[Info] Test data generated and saved as 'test_income_data.csv'.")

# 분석 실행
def run_analysis():
    # 테스트 데이터 생성
    generate_test_data()

    # 분석할 회사 목록
    selected_companies = ["A사", "B사"]

    # 데이터 경로 설정 (테스트 파일 경로 사용)
    income_data_path = r"C:\Users\CAD09\Downloads\miniproject-Analysis-Graph-1.0.4\datasets\IS"
    balance_data_path =   r"C:\Users\CAD09\Downloads\miniproject-Analysis-Graph-1.0.4\datasets\BS"  # 현재 디렉토리

    # 데이터 준비
    income_data, balance_data = prepare_data_for_analysis(balance_data_path, income_data_path, selected_companies)

    # 분석 결과 생성 (단순 매출 데이터 추출)
    analysis_results = []
    for company in selected_companies:
        for data_item in ["ifrs-full_Revenue", "ifrs-full_Profit"]:
            filtered_data = income_data[
                (income_data["회사명"] == company) & (income_data["항목코드"] == data_item)
            ]
            for _, row in filtered_data.iterrows():
                analysis_results.append({
                    "회사명": company,
                    "항목명": data_item,
                    "2019": row["당기"] if row["연도"] == 2019 else None,
                    "2020": row["당기"] if row["연도"] == 2020 else None
                })

    # Excel 저장 로직 실행
    formatter = ExcelFormatter(output_file="financial_analysis_output.xlsx")
    formatter.set_zoom()
    formatter.set_dimensions()
    formatter.fill_background()

    # 테두리, 색상 및 텍스트 설정
    specified_ranges = ["B4:I10"]
    merge_and_set = [
        ("B4:C5", "회사명"), ("D4:D5", "2019년"), ("E4:E5", "2020년")
    ]
    formatter.apply_border(specified_ranges)
    formatter.merge_cells_and_set_text(merge_and_set)

    # 데이터 작성
    row_mapping = {
        "ifrs-full_Revenue": 6,
        "ifrs-full_Profit": 7
    }
    for result in analysis_results:
        company_row = row_mapping.get(result["항목명"])
        if company_row:
            formatter.ws[f"D{company_row}"].value = result.get("2019", "")
            formatter.ws[f"E{company_row}"].value = result.get("2020", "")

    formatter.save_file()

    # Plotter 예제 실행
    plotter = Plotter("A사", income_data)
    fig = plotter.create_graph("ifrs-full_Revenue", "매출")
    plotter.save_graph_as_png(fig, "ifrs-full_Revenue", "손익계산서")

# 테스트 실행
if __name__ == "__main__":
    run_analysis()
